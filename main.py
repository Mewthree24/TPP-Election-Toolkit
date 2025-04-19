import streamlit as st
import json
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment
from io import BytesIO
from collections import defaultdict
import os
import streamlit.components.v1 as components
import re
from streamlit_javascript import st_javascript

st.set_page_config(page_title="TPP Election Toolkit", layout="wide")

# === Color Generation Functions ===
def normalize_county_id(name, state=None):
    name = name.lower().strip()

    # Remove generic suffixes
    name = (name.replace(" census area", "")
            .replace(" city and borough", "")
            .replace(" municipality", "")
            .replace(" borough", "")
            .replace(" county", "")
            .replace(".", "")
            .replace("'", "")
            .replace("-", "_")
            .replace(" ", "_"))

    # Fix common abbreviations
    name = name.replace("st_", "st").replace("ste_", "ste")

    # Return normalized name
    return name

def build_state_color_map(df, dem_colors, rep_colors, ind_colors):
    color_map = {}
    rating_to_color = {
        f"{strength} Democratic": dem_colors.get(strength, "#cccccc")
        for strength in ["Safe", "Likely", "Lean", "Tilt"]
    }
    rating_to_color.update({
        f"{strength} Republican": rep_colors.get(strength, "#cccccc")
        for strength in ["Safe", "Likely", "Lean", "Tilt"]
    })
    rating_to_color.update({
        f"{strength} Independent": ind_colors.get(strength, "#cccccc")
        for strength in ["Safe", "Likely", "Lean", "Tilt"]
    })

    state_name_to_code = {
        "Alabama": "AL", "Alaska": "AK", "Arizona": "AZ", "Arkansas": "AR",
        "California": "CA", "Colorado": "CO", "Connecticut": "CT", "Delaware": "DE",
        "Florida": "FL", "Georgia": "GA", "Hawaii": "HI", "Idaho": "ID",
        "Illinois": "IL", "Indiana": "IN", "Iowa": "IA", "Kansas": "KS",
        "Kentucky": "KY", "Louisiana": "LA", "Maine": "ME", "Maryland": "MD",
        "Massachusetts": "MA", "Michigan": "MI", "Minnesota": "MN", "Mississippi": "MS",
        "Missouri": "MO", "Montana": "MT", "Nebraska": "NE", "Nevada": "NV",
        "New Hampshire": "NH", "New Jersey": "NJ", "New Mexico": "NM", "New York": "NY",
        "North Carolina": "NC", "North Dakota": "ND", "Ohio": "OH", "Oklahoma": "OK",
        "Oregon": "OR", "Pennsylvania": "PA", "Rhode Island": "RI", "South Carolina": "SC",
        "South Dakota": "SD", "Tennessee": "TN", "Texas": "TX", "Utah": "UT",
        "Vermont": "VT", "Virginia": "VA", "Washington": "WA", "West Virginia": "WV",
        "Wisconsin": "WI", "Wyoming": "WY", "District of Columbia": "DC"
    }

    for _, row in df.iterrows():
        state = row.get("State", "")
        rating = row.get("Rating", "")

        if pd.isna(state) or pd.isna(rating):
            continue

        # Normalize state to 2-letter code
        normalized = state.strip()
        if len(normalized) > 2:  # convert full name to code
            normalized = state_name_to_code.get(normalized, "")
        else:
            normalized = normalized.upper()

        color = rating_to_color.get(rating.strip(), "#cccccc")
        color_map[normalized] = color

    return color_map

def build_county_color_map(df, dem_colors, rep_colors, ind_colors):
    color_map = {}
    rating_to_color = {
        f"{strength} Democratic": dem_colors.get(strength, "#cccccc")
        for strength in ["Safe", "Likely", "Lean", "Tilt"]
    }
    rating_to_color.update({
        f"{strength} Republican": rep_colors.get(strength, "#cccccc")
        for strength in ["Safe", "Likely", "Lean", "Tilt"]
    })
    rating_to_color.update({
        f"{strength} Independent": ind_colors.get(strength, "#cccccc")
        for strength in ["Safe", "Likely", "Lean", "Tilt"]
    })

    for _, row in df.iterrows():
        county = row.get("County", "")
        rating = row.get("Rating", "")

        if pd.isna(county) or pd.isna(rating):
            continue

        county = str(county).strip()
        rating = str(rating).strip()

        if not county or not rating:
            continue

        # Properly normalize county ID
        county_id = normalize_county_id(county, state_code)
        color = rating_to_color.get(rating, "#cccccc")
        color_map[county_id] = color

    return color_map

def apply_state_colors_to_svg(svg_text, color_map):
    def replace_fill(match):
        tag = match.group(0)
        tag_type = match.group(1)
        tag_id = match.group(2)

        # SVG uses 2-letter state codes like "NY" 
        normalized_id = tag_id.strip().upper()
        color = color_map.get(normalized_id)

        if color:
            if 'style=' in tag:
                tag = re.sub(r'fill:[^;"]+', f'fill:{color}', tag)
            else:
                tag = tag.replace('>', f' style="fill:{color}">')
        return tag

    # Match all valid SVG shape tags
    return re.sub(r'<(path|g|rect|polygon|polyline|circle)[^>]*id="([^"]+)"[^>]*>', replace_fill, svg_text)

def apply_county_colors_to_svg(svg_text, color_map, state_code):
    def replace_fill(match):
        tag = match.group(0)
        tag_type = match.group(1)
        tag_id = match.group(2)

        normalized_id = tag_id.strip().lower()

        # Normalize county formatting
        normalized_id = normalized_id.replace("-", "_").replace(" ", "_").replace("county", "").replace(".", "").replace("'", "")
        normalized_id = normalized_id.replace("st_", "st").replace("ste_", "ste").strip("_")

        color = color_map.get(normalized_id)
        if color:
            if 'style=' in tag:
                tag = re.sub(r'fill:[^;"]+', f'fill:{color}', tag)
            else:
                tag = tag.replace('>', f' style="fill:{color}">')
        return tag

    # Apply fill color to matching SVG elements
    return re.sub(r'<(path|g|rect|polygon|polyline|circle)[^>]*id="([^"]+)"[^>]*>', replace_fill, svg_text)


def display_national_map(election_type):
    """Helper function to display national maps for President/Senate/Governor"""
    map_file = {
        "President": "presidential.svg",
        "Senate": "states.svg",
        "Governor": "states.svg"
    }.get(election_type)

    if map_file:
        path = os.path.join("SVG", map_file)
        if os.path.exists(path):
            render_svg_file(path, title=f"🗺️ {election_type} National Map")
        else:
            st.warning(f"No national map found for {election_type}")

def render_svg_file(svg_path: str, title: str = None, df_display=None, dem_colors=None, rep_colors=None, ind_colors=None):
    import streamlit.components.v1 as components
    import base64
    import os

    try:
        with open(svg_path, "r", encoding="utf-8") as f:
            svg_data = f.read()

        if title:
            st.subheader(title)

        # Apply coloring if we have display data and color schemes
        if df_display is not None and dem_colors and rep_colors and ind_colors:
            # Choose map type based on filename
            if "presidential" in svg_path or "states" in svg_path:
                color_map = build_state_color_map(df_display, dem_colors, rep_colors, ind_colors)
                svg_data = apply_state_colors_to_svg(svg_data, color_map)
            else:
                color_map = build_county_color_map(df_display, dem_colors, rep_colors, ind_colors)
                svg_data = apply_county_colors_to_svg(svg_data, color_map, state_code)

        encoded = base64.b64encode(svg_data.encode()).decode()

        # === Force proper aspect ratio rendering (like image) and add message handler ===
        # Fix aspect ratio by injecting correct SVG attributes
        svg_display = re.sub(
            r'<svg([^>]*)>',
            r'<svg\1 preserveAspectRatio="xMidYMid meet">',
            svg_data
        )

        # Inject viewBox if not present and ensure preserveAspectRatio is set
        if not re.search(r'viewBox=', svg_display):
            svg_display = re.sub(r'<svg', '<svg viewBox="0 0 1000 600"', svg_display)

        # Create base64 of modified SVG
        encoded = base64.b64encode(svg_display.encode()).decode()

        components.html(
            f"""
            <div style="width: 100%; display: flex; justify-content: center; align-items: center;">
                <div style="width: 100%; max-width: 1200px; margin: 0 auto;">
                    <div style="position: relative; width: 100%; padding-bottom: 75%;">
                        <object data="data:image/svg+xml;base64,{encoded}"
                                type="image/svg+xml"
                                style="position: absolute; top: 0; left: 0; width: 100%; height: 100%; object-fit: contain;">
                        </object>
                    </div>
                </div>
            </div>
            """,
            height=700,
            scrolling=False
        )

        # Add download button for colored map
        st.download_button(
            label="📥 Download SVG Map",
            data=svg_display.encode("utf-8"),
            file_name=os.path.basename(svg_path).replace(".svg", "_colored.svg"),
            mime="image/svg+xml"
        )

        st.success(f"🗺️ Displaying: {os.path.basename(svg_path)}")

    except Exception as e:
        st.error(f"⚠️ Failed to render SVG: {e}")

# === Map Generation ===
svg_folder_path = os.path.join(os.getcwd(), "SVG")
svg_files = [f for f in os.listdir(svg_folder_path) if f.endswith(".svg")]

# Categorize them
pres_svg = "presidential.svg" in svg_files
states_svg = "states.svg" in svg_files
state_svgs = sorted([f for f in svg_files if f not in ("presidential.svg", "states.svg")])

# Initialize session
if "election_data" not in st.session_state:
    st.session_state["election_data"] = {}

# Initialize margin thresholds
if "tilt_max" not in st.session_state:
    st.session_state["tilt_max"] = 3
if "lean_max" not in st.session_state:
    st.session_state["lean_max"] = 7
if "likely_max" not in st.session_state:
    st.session_state["likely_max"] = 12

# Initialize color settings
if "color_settings" not in st.session_state:
    st.session_state["color_settings"] = {
      "Democratic": {
        "Tilt": "#949BB3",
        "Lean": "#8AAFFF",
        "Likely": "#577CCC",
        "Safe": "#1C408C"
      },
      "Republican": {
        "Tilt": "#CF8980",
        "Lean": "#FF8B98",
        "Likely": "#FF5865",
        "Safe": "#BF1D29"
      },
      "Independent": {
        "Tilt": "#FEF4B4",
        "Lean": "#FED463",
        "Likely": "#FE9929",
        "Safe": "#CC4C02"
      }
    }

st.title("🗳️ TPP Election Toolkit")

# Upload file
uploaded_file = st.file_uploader("Upload your savefile", type=["json"])

if uploaded_file:
    try:
        raw_data = uploaded_file.read()
        data = json.loads(raw_data)

        wanted_keys = [
            "electNightSB", "electNightCC", "electNightM",
            "electNightStH", "electNightStS", "electNightG",
            "electNightUSH", "electNightUSS", "electNightP"
        ]

        extracted_data = {k: data[k] for k in wanted_keys if k in data}
        st.session_state["election_data"] = extracted_data

        st.success("Election data extracted successfully.")

    except Exception as e:
        st.error(f"Failed to load file: {e}")

# Initialize session state for selected state
if "selected_state" not in st.session_state:
    st.session_state.selected_state = "National View"

if st.session_state["election_data"]:

    state_code_to_name = {
        "AL": "Alabama", "AK": "Alaska", "AZ": "Arizona", "AR": "Arkansas",
        "CA": "California", "CO": "Colorado", "CT": "Connecticut", "DE": "Delaware",
        "FL": "Florida", "GA": "Georgia", "HI": "Hawaii", "ID": "Idaho",
        "IL": "Illinois", "IN": "Indiana", "IA": "Iowa", "KS": "Kansas",
        "KY": "Kentucky", "LA": "Louisiana", "ME": "Maine", "MD": "Maryland",
        "MA": "Massachusetts", "MI": "Michigan", "MN": "Minnesota", "MS": "Mississippi",
        "MO": "Missouri", "MT": "Montana", "NE": "Nebraska", "NV": "Nevada",
        "NH": "New Hampshire", "NJ": "New Jersey", "NM": "New Mexico", "NY": "New York",
        "NC": "North Carolina", "ND": "North Dakota", "OH": "Ohio", "OK": "Oklahoma",
        "OR": "Oregon", "PA": "Pennsylvania", "RI": "Rhode Island", "SC": "South Carolina",
        "SD": "South Dakota", "TN": "Tennessee", "TX": "Texas", "UT": "Utah",
        "VT": "Vermont", "VA": "Virginia", "WA": "Washington", "WV": "West Virginia",
        "WI": "Wisconsin", "WY": "Wyoming", "DC": "District of Columbia"
    }

    election_types = {
        "President": "electNightP",
        "Senate": "electNightUSS",
        "Governor": "electNightG",
        "U.S. House": "electNightUSH",
        "State House": "electNightStH",
        "State Senate": "electNightStS"
    }

    available_election_types = [etype for etype, ekey in election_types.items() if ekey in st.session_state["election_data"]]

    if available_election_types:
        selected_election_type = st.selectbox("Select Election Type", available_election_types)
        
        # Initialize state selection if not present
        if "selected_state" not in st.session_state:
            st.session_state.selected_state = "National View"

        # Check for clicked state from localStorage
        key_base = f"{selected_election_type}_{hash(str(st.session_state.get('selected_state', '')))}"
        clicked_state = st_javascript("""
            const stored = localStorage.getItem("clicked_state");
            if (stored) {
                localStorage.removeItem("clicked_state");
                return stored;
            }
            return null;
        """, key=f"check_clicked_state_{key_base}")

        if clicked_state:
            st.session_state.selected_state = clicked_state
            st.experimental_rerun()
        election_key = election_types[selected_election_type]
        election_data = st.session_state["election_data"][election_key]

        def assign_rating(margin, winner, tilt_max, lean_max, likely_max):
            if margin <= tilt_max:
                level = "Tilt"
            elif margin <= lean_max:
                level = "Lean"
            elif margin <= likely_max:
                level = "Likely"
            else:
                level = "Safe"
            return f"{level} {winner}"

        def update_df_with_custom_ratings(df, tilt_max=None, lean_max=None, likely_max=None):
            df = df.copy()
            if "Rating" in df.columns and "Margin %" in df.columns:
                # Use session state values if not provided
                tilt = tilt_max if tilt_max is not None else st.session_state.get("tilt_max", 3)
                lean = lean_max if lean_max is not None else st.session_state.get("lean_max", 7)
                likely = likely_max if likely_max is not None else st.session_state.get("likely_max", 12)

                df["Rating"] = df.apply(
                    lambda row: assign_rating(
                        abs(float(str(row["Margin %"]).strip("%"))),
                        row["Rating"].split()[-1] if isinstance(row["Rating"], str) else "",
                        tilt, lean, likely
                    )
                    if pd.notna(row.get("Rating")) and "%" in str(row["Margin %"])
                    else row.get("Rating", ""),
                    axis=1
                )
            return df

        entries_to_convert = []

        # === U.S. House National View Spreadsheet Generator ===
        if selected_election_type == "U.S. House":
            # Margin thresholds for House - now using session state
            tilt_max = st.slider("Tilt Margin Max (%)", 1, 5, 3, key="house_tilt")
            lean_max = st.slider("Lean Margin Max (%)", 5, 10, 7, key="house_lean")
            likely_max = st.slider("Likely Margin Max (%)", 10, 20, 12, key="house_likely")

            from collections import defaultdict
            wb = Workbook()
            ws = wb.active
            ws.title = "U.S. House National View"

            entries = election_data.get("elections", [])
            party_labels = {"D": "Democratic", "R": "Republican", "I": "Independent"}
            party_order = ["D", "R", "I"]
            seats_won = {party: 0 for party in party_order} # Added to track seats won

            # === Header Rows ===
            ws.cell(row=2, column=1, value="State")
            ws.cell(row=2, column=2, value="District")
            col = 3
            for party in party_order:
                ws.cell(row=1, column=col, value=party_labels[party])
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
                ws.cell(row=2, column=col, value="Candidate")
                ws.cell(row=2, column=col + 1, value="#")
                ws.cell(row=2, column=col + 2, value="%")
                col += 3

            ws.cell(row=1, column=col, value="Margins & Rating")
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
            ws.cell(row=2, column=col, value="Margin #")
            ws.cell(row=2, column=col + 1, value="Margin %")
            ws.cell(row=2, column=col + 2, value="Total Vote")
            ws.cell(row=2, column=col + 3, value="Rating")

            for r in range(1, 3):
                for c in range(1, col + 4):
                    cell = ws.cell(row=r, column=c)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

            row_idx = 3
            totals = {party: 0 for party in party_order}
            grand_total = 0

            for entry in entries:
                state = entry.get("state", "??")
                district = entry.get("district", "?")
                ws.cell(row=row_idx, column=1, value=state)
                ws.cell(row=row_idx, column=2, value=district)

                # Group candidates by party
                party_groups = defaultdict(list)
                for c in entry.get("cands", []):
                    party_groups[c["party"]].append(c)

                # Determine winner
                all_cands = sorted(entry.get("cands", []), key=lambda x: x["votes"], reverse=True)
                winner = all_cands[0]["name"] if all_cands else None
                winner_party = all_cands[0]["party"] if all_cands else None

                # Prepare vote summary by party
                party_votes = {}
                party_names = {}
                total_vote = sum(c["votes"] for c in entry.get("cands", []))

                for party in party_order:
                    candidates = sorted(party_groups.get(party, []), key=lambda x: x["votes"], reverse=True)
                    if not candidates:
                        party_names[party] = ""
                        party_votes[party] = 0
                        continue

                    if len(candidates) == 1:
                        party_names[party] = candidates[0]["name"]
                        party_votes[party] = candidates[0]["votes"]
                    else:
                        if candidates[0]["name"] == winner:
                            combined = sum(c["votes"] for c in candidates)
                            party_names[party] = candidates[0]["name"]
                            party_votes[party] = combined
                        else:
                            party_names[party] = candidates[0]["name"]
                            party_votes[party] = candidates[0]["votes"]
                            # Move lowest-vote candidate to Independent
                            lowest = candidates[-1]
                            # Safely initialize Independent party
                            if "I" not in party_votes:
                                party_votes["I"] = 0
                            if "I" not in party_names:
                                party_names["I"] = ""
                            party_names["I"] = lowest["name"]
                            party_votes["I"] += lowest["votes"]

                col_idx = 3
                for party in party_order:
                    name = party_names.get(party, "")
                    votes = int(round(party_votes.get(party, 0)))
                    pct = round(votes / total_vote * 100, 2) if total_vote else 0

                    ws.cell(row=row_idx, column=col_idx, value=name)
                    ws.cell(row=row_idx, column=col_idx + 1, value=f"{votes:,}")
                    ws.cell(row=row_idx, column=col_idx + 2, value=f"{pct:.2f}%")

                    totals[party] += votes
                    col_idx += 3

                sorted_votes = sorted(party_votes.items(), key=lambda x: x[1], reverse=True)
                margin = int(round(sorted_votes[0][1] - (sorted_votes[1][1] if len(sorted_votes) > 1 else 0)))
                margin_pct = round(margin / total_vote * 100, 2) if total_vote else 0
                rating = assign_rating(margin_pct, party_labels.get(sorted_votes[0][0], sorted_votes[0][0]), tilt_max, lean_max, likely_max)

                ws.cell(row=row_idx, column=col_idx, value=f"{margin:,}")
                ws.cell(row=row_idx, column=col_idx + 1, value=f"{margin_pct:.2f}%")
                ws.cell(row=row_idx, column=col_idx + 2, value=f"{int(round(total_vote)):,}")
                ws.cell(row=row_idx, column=col_idx + 3, value=rating)

                if winner_party:
                    seats_won[winner_party] +=1 #update seats won

                grand_total += total_vote
                row_idx += 1

            # === Totals Row ===
            ws.cell(row=row_idx, column=1, value="TOTALS")
            ws.cell(row=row_idx, column=2, value="")
            col_idx = 3
            for party in party_order:
                total = totals[party]
                pct = round(total / grand_total * 100, 2) if grand_total else 0
                ws.cell(row=row_idx, column=col_idx, value=f"{seats_won[party]} seats")
                ws.cell(row=row_idx, column=col_idx + 1, value=f"{total:,}")
                ws.cell(row=row_idx, column=col_idx + 2, value=f"{pct:.2f}%")
                col_idx += 3

            # Calculate margin for totals row
            sorted_totals = sorted(totals.items(), key=lambda x: x[1], reverse=True)
            margin_total = sorted_totals[0][1] - (sorted_totals[1][1] if len(sorted_totals) > 1 else 0)
            margin_pct_total = round(margin_total / grand_total * 100, 2) if grand_total else 0
            winner_party = sorted_totals[0][0]
            rating = assign_rating(margin_pct_total, party_labels.get(winner_party, winner_party), tilt_max, lean_max, likely_max)

            ws.cell(row=row_idx, column=col_idx, value=f"{margin_total:,}")
            ws.cell(row=row_idx, column=col_idx + 1, value=f"{margin_pct_total:.2f}%")
            ws.cell(row=row_idx, column=col_idx + 2, value=f"{int(round(grand_total)):,}")
            ws.cell(row=row_idx, column=col_idx + 3, value=rating)

            for c in range(1, col_idx + 4):
                ws.cell(row=row_idx, column=c).font = Font(bold=True)

            # === Streamlit Display ===
            from io import BytesIO
            st.subheader("🧾 U.S. House National View")
            file_stream = BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)

            # Convert worksheet to displayable rows
            excel_rows = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                excel_rows.append(list(row))

            header_row = []
            data_rows = []

            if len(excel_rows) >= 2:
                row1 = excel_rows[0]
                row2 = excel_rows[1]
                used_names = {}

                for col1, col2 in zip(row1, row2):
                    if col1 and col2:
                        label = f"{col1} - {col2}"
                    elif col1:
                        label = str(col1)
                    elif col2:
                        label = str(col2)
                    else:
                        label = "Unnamed"

                    if label in used_names:
                        count = used_names[label] + 1
                        used_names[label] = count
                        label = f"{label} ({count})"
                    else:
                        used_names[label] = 1

                    header_row.append(label)

                data_rows = excel_rows[2:]

            # Show dataframe
            if header_row and data_rows:
                df_display = pd.DataFrame(data_rows, columns=header_row)
                # Convert all numeric-like strings to numeric values
                df_display = df_display.apply(pd.to_numeric, errors='ignore')
                # Apply custom ratings based on session state thresholds
                df_display = update_df_with_custom_ratings(
                    df_display,
                    tilt_max,
                    lean_max,
                    likely_max
                )
                st.dataframe(df_display, use_container_width=True)

            # Download button
            st.download_button(
                label="📥 Download House Spreadsheet",
                data=file_stream,
                file_name="House_National_View.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key="house_national_view"
            )

        elif selected_election_type in ["President", "Senate", "Governor"]:
            available_states = [entry.get("state") for entry in election_data.get("elections", [])]
            available_states = sorted(set(available_states))
            state_options = ["National View"] + [state_code_to_name.get(code, code) for code in available_states]
            selected_state = st.selectbox(
                "Select State",
                state_options,
                index=state_options.index(st.session_state.get("selected_state", "National View")),
                key="select_state_box"
            )

            # === Show UI Controls Before Any Map ===
            if selected_election_type in ["President", "Senate", "Governor"] and selected_state != "None":
                st.markdown("### 🎯 Margin Thresholds")
                tilt_max = st.slider("Tilt Margin Max (%)", 1, 5, 3, key="slider_tilt")
                lean_max = st.slider("Lean Margin Max (%)", 5, 10, 7, key="slider_lean")
                likely_max = st.slider("Likely Margin Max (%)", 10, 20, 12, key="slider_likely")

                st.markdown("### 🎨 Color Customizer")
                col1, col2, col3 = st.columns(3)

                # Add save/load UI in a container to keep it compact
                with st.container():
                    save_col, load_col = st.columns(2)

                    with save_col:
                        if st.button("💾 Save Colors"):
                            json_str = json.dumps(st.session_state["color_settings"], indent=2)
                            st.download_button(
                                label="📥 Download Settings",
                                data=json_str,
                                file_name="color_settings.json",
                                mime="application/json"
                            )

                    with load_col:
                        uploaded_file = st.file_uploader("📤 Load Colors", type=["json"])
                        if uploaded_file is not None:
                            try:
                                uploaded_colors = json.load(uploaded_file)
                                if all(party in uploaded_colors for party in ["Democratic", "Republican", "Independent"]):
                                    st.session_state["color_settings"] = uploaded_colors
                                    st.success("✅ Colors loaded!")
                                else:
                                    st.error("⚠️ Invalid color file format")
                            except Exception as e:
                                st.error(f"⚠️ Failed to load colors: {e}")

                # Create a unique identifier based on location in code
                color_ui_id = f"national_view_{hash(selected_state + selected_election_type)}"

                # Shared color levels for all parties
                color_levels = ["Tilt", "Lean", "Likely", "Safe"]
                color_mapping = {
                    "Democratic": {"label": "Dem", "colors": {}},
                    "Republican": {"label": "Rep", "colors": {}},
                    "Independent": {"label": "Ind", "colors": {}}
                }

                with col1:
                    st.markdown("**Democratic Shades**")
                    for level in color_levels:
                        color_key = f"dem_{level.lower()}"
                        st.session_state.color_settings["Democratic"][level] = st.color_picker(
                            f"{level} Dem",
                            value=st.session_state.color_settings["Democratic"][level],
                            key=color_key
                        )
                        color_mapping["Democratic"]["colors"][level] = st.session_state.color_settings["Democratic"][level]

                with col2:
                    st.markdown("**Republican Shades**")
                    for level in color_levels:
                        color_key = f"rep_{level.lower()}"
                        st.session_state.color_settings["Republican"][level] = st.color_picker(
                            f"{level} Rep",
                            value=st.session_state.color_settings["Republican"][level],
                            key=color_key
                        )
                        color_mapping["Republican"]["colors"][level] = st.session_state.color_settings["Republican"][level]

                with col3:
                    st.markdown("**Independent Shades**")
                    for level in color_levels:
                        color_key = f"ind_{level.lower()}"
                        st.session_state.color_settings["Independent"][level] = st.color_picker(
                            f"{level} Ind",
                            value=st.session_state.color_settings["Independent"][level],
                            key=color_key
                        )
                        color_mapping["Independent"]["colors"][level] = st.session_state.color_settings["Independent"][level]

                # Set the color variables for map rendering
                dem_colors = color_mapping["Democratic"]["colors"]
                rep_colors = color_mapping["Republican"]["colors"]
                ind_colors = color_mapping["Independent"]["colors"]

            if selected_state != "National View":
                state_code = next((code for code, name in state_code_to_name.items() if name == selected_state), None)
                if state_code:
                    state_entries = [entry for entry in election_data.get("elections", []) if entry.get("state") == state_code]
                    if state_entries:
                        state_entry = state_entries[0]
                        counties = state_entry.get("counties", [])
                        spreadsheet_rows = []

                        for county in counties:
                            county_name = county.get("name", "Unknown County")
                            for cand in county.get("cands", []):
                                spreadsheet_rows.append({
                                    "State": state_code,
                                    "County": county_name,
                                    "Candidate": cand.get("name", ""),
                                    "Party": cand.get("party", ""),
                                    "Votes": cand.get("votes", 0),
                                    "Incumbent": cand.get("incumbent", False),
                                    "Caucus": cand.get("caucus", ""),
                                })

                        if spreadsheet_rows:
                            df = pd.DataFrame(spreadsheet_rows)

                            wb = Workbook()
                            ws = wb.active
                            ws.title = f"{state_code} County Results"

                            # === Create party blocks dynamically from candidates ===
                            party_labels = {"D": "Democratic", "R": "Republican", "I": "Independent"}

                            candidates = state_entry.get("cands", [])
                            party_to_candidates = {}

                            for cand in candidates:
                                party = cand["party"]
                                if party not in party_to_candidates:
                                    party_to_candidates[party] = []
                                party_to_candidates[party].append(cand["name"])

                            party_order = list(party_to_candidates.keys())

                            # === Header rows ===
                            ws.cell(row=2, column=1, value="County")
                            col = 2

                            for party in party_order:
                                full_party_name = party_labels.get(party, party)
                                candidate_names = party_to_candidates[party]

                                span = len(candidate_names) * 2
                                if span > 0:
                                    ws.cell(row=1, column=col, value=full_party_name)
                                    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + span - 1)

                                    for cand in candidate_names:
                                        ws.cell(row=2, column=col, value=cand)
                                        ws.cell(row=2, column=col + 1, value="%")
                                        col += 2

                            ws.cell(row=1, column=col, value="Margins & Rating")
                            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
                            ws.cell(row=2, column=col, value="Margin #")
                            ws.cell(row=2, column=col + 1, value="Margin %")
                            ws.cell(row=2, column=col + 2, value="Total Vote")
                            ws.cell(row=2, column=col + 3, value="Rating")

                            # Format headers
                            for r in range(1, 3):
                                for c in range(1, col + 4):
                                    cell = ws.cell(row=r, column=c)
                                    cell.font = Font(bold=True)
                                    cell.alignment = Alignment(horizontal="center", vertical="center")

                            # === Data rows ===
                            row_idx = 3
                            ordered_candidates = [(party, name) for party in party_order for name in party_to_candidates[party]]
                            candidate_totals = {cand_name: 0 for _, cand_name in ordered_candidates}

                            for county in counties:
                                clean_name = county.get("name", "Unknown County").replace(" County", "").title()
                                ws.cell(row=row_idx, column=1, value=clean_name)

                                vote_map = {c["name"]: round(c["votes"], 2) for c in county.get("cands", [])}
                                total_vote = sum(vote_map.values())
                                vote_values = []

                                col = 2
                                for _, candidate_name in ordered_candidates:
                                    v = int(round(vote_map.get(candidate_name, 0)))
                                    pct = round(v / total_vote * 100, 2) if total_vote else 0
                                    ws.cell(row=row_idx, column=col, value="{:,}".format(v))
                                    ws.cell(row=row_idx, column=col + 1, value="{:.2f}%".format(pct))
                                    candidate_totals[candidate_name] += v
                                    vote_values.append((candidate_name, v))
                                    col += 2

                                vote_values.sort(key=lambda x: x[1], reverse=True)
                                margin = vote_values[0][1] - vote_values[1][1] if len(vote_values) > 1 else vote_values[0][1]
                                margin_pct = round(margin / total_vote * 100, 2) if total_vote else 0
                                winner = vote_values[0][0]
                                winner_party = next((c["party"] for c in candidates if c["name"] == winner), "?")
                                rating = assign_rating(margin_pct, party_labels.get(winner_party, winner_party), tilt_max, lean_max, likely_max)

                                ws.cell(row=row_idx, column=col, value="{:,}".format(margin))
                                ws.cell(row=row_idx, column=col + 1, value="{:.2f}%".format(margin_pct))
                                ws.cell(row=row_idx, column=col + 2, value="{:,}".format(int(round(total_vote))))
                                ws.cell(row=row_idx, column=col + 3, value=rating)
                                row_idx += 1

                            # === Totals row ===
                            ws.cell(row=row_idx, column=1, value="TOTALS")
                            grand_total = sum(candidate_totals.values())
                            col = 2
                            for _, cand_name in ordered_candidates:
                                v = int(round(candidate_totals[cand_name]))
                                pct = round(v / grand_total * 100, 2) if grand_total else 0
                                ws.cell(row=row_idx, column=col, value="{:,}".format(v))
                                ws.cell(row=row_idx, column=col + 1, value="{:.2f}%".format(pct))
                                col += 2

                            # Margin/Rating
                            sorted_totals = sorted(candidate_totals.items(), key=lambda x: x[1], reverse=True)
                            top = sorted_totals[0][1]
                            second = sorted_totals[1][1] if len(sorted_totals) > 1 else 0
                            margin = top - second
                            margin_pct = round(margin / grand_total * 100, 2) if grand_total else 0
                            winner_party = next((c["party"] for c in candidates if c["name"] == sorted_totals[0][0]), "?")
                            rating = assign_rating(margin_pct, party_labels.get(winner_party, winner_party), tilt_max, lean_max, likely_max)

                            ws.cell(row=row_idx, column=col, value="{:,}".format(margin))
                            ws.cell(row=row_idx, column=col + 1, value="{:.2f}%".format(margin_pct))
                            ws.cell(row=row_idx, column=col + 2, value="{:,}".format(int(round(grand_total))))
                            ws.cell(row=row_idx, column=col + 3, value=rating)

                        from openpyxl.styles import Font

                        # Bold Totals row
                        for col_idx in range(1, col + 4):
                            ws.cell(row=row_idx, column=col_idx).font = Font(bold=True)

                        # Save file to memory
                        file_stream = BytesIO()
                        wb.save(file_stream)
                        file_stream.seek(0)

                        # === DISPLAY FORMATTED PREVIEW IN STREAMLIT ===
                        # Collect rows from worksheet
                        excel_rows = []
                        for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                            excel_rows.append(list(row))

                            # Merge row 1 and 2 into a single header string with uniqueness
                        header_row = []
                        if len(excel_rows) >= 2:
                            row1 = excel_rows[0]
                            row2 = excel_rows[1]
                            used_names = {}

                            for col1, col2 in zip(row1, row2):
                                    if col1 and col2:
                                        label = f"{col1} - {col2}"
                                    elif col1:
                                        label = str(col1)
                                    elif col2:
                                        label = str(col2)
                                    else:
                                        label = "Unnamed"

                                    # Ensure uniqueness
                                    if label in used_names:
                                        count = used_names[label] + 1
                                        used_names[label] = count
                                        label = f"{label} ({count})"
                                    else:
                                        used_names[label] = 1

                                    header_row.append(label)

                        # Use remaining rows as data
                        data_rows = excel_rows[2:]

                        # Display in Streamlit
                        df_display = pd.DataFrame(data_rows, columns=header_row)
                        # Convert all numeric-like strings to numeric values
                        df_display = df_display.apply(pd.to_numeric, errors='ignore')
                        st.subheader(f"🧾 {selected_state} County-Level Results")
                        st.dataframe(df_display, use_container_width=True)

                        # Create download button (one time only)
                        st.download_button(
                            label="📥 Download County-Level Spreadsheet", 
                            data=file_stream,
                            file_name = f"{state_code}_{selected_election_type}_County_Results.xlsx",
                            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            key=f"county_download_{state_code}"
                        )

                        # Show county-level map
                        svg_filename = f"{state_code.lower()}.svg"
                        svg_path = os.path.join("SVG", svg_filename)
                        if os.path.exists(svg_path):
                            # Extract County and Rating from displayed df
                            coloring_df = df_display[["County", "Rating"]].copy()
                            render_svg_file(svg_path, title="🗺️ County-Level Map", df_display=coloring_df, dem_colors=dem_colors, rep_colors=rep_colors, ind_colors=ind_colors)
                        else:
                            st.warning(f"❌ No county-level map found for {state_code}")
                    else:
                        st.warning("No county-level data found.")
                else:
                    st.warning("Selected state not found.")
            else:
                # === Presidential National View Spreadsheet ===
                entries_to_convert = election_data.get("elections", [])

                if selected_election_type == "President":
                    wb = Workbook()
                    ws = wb.active
                    ws.title = "Presidential National View"

                    candidates = []
                    party_labels = {"D": "Democratic", "R": "Republican", "I": "Independent"}
                    party_to_candidate = {}

                    # Extract candidate names and parties
                    if entries_to_convert:
                        first_entry = entries_to_convert[0]
                        for cand in first_entry.get("cands", []):
                            party = cand["party"]
                            name = cand["name"]
                            party_to_candidate[party] = name

                        candidate_parties = list(party_to_candidate.keys())

                        # === Header rows ===
                        ws.cell(row=2, column=1, value="State")
                        ws.cell(row=2, column=2, value="Electoral Votes")
                        col = 3

                        for party in candidate_parties:
                            full_party = party_labels.get(party, party)
                            candidate = party_to_candidate[party]

                            ws.cell(row=1, column=col, value=full_party)
                            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)

                            ws.cell(row=2, column=col, value=candidate)  # Raw votes
                            ws.cell(row=2, column=col + 1, value="%")
                            ws.cell(row=2, column=col + 2, value="#")
                            col += 3

                        ws.cell(row=1, column=col, value="Margins & Rating")
                        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
                        ws.cell(row=2, column=col, value="Margin #")
                        ws.cell(row=2, column=col + 1, value="Margin %")
                        ws.cell(row=2, column=col + 2, value="Total Vote")
                        ws.cell(row=2, column=col + 3, value="Rating")

                        for r in range(1, 3):
                            for c in range(1, col + 4):
                                cell = ws.cell(row=r, column=c)
                                cell.font = Font(bold=True)
                                cell.alignment = Alignment(horizontal="center", vertical="center")

                        # === Data Rows ===
                        row_idx = 3
                        total_votes = {p: 0 for p in candidate_parties}
                        electoral_totals = {p: 0 for p in candidate_parties}

                        all_states = {e["state"]: e for e in entries_to_convert}

                        for state_code in state_code_to_name:
                            entry = all_states.get(state_code)
                            if not entry:
                                continue

                            ws.cell(row=row_idx, column=1, value=state_code_to_name[state_code])
                            state_votes = {c["name"]: c["votes"] for c in entry["cands"]}
                            party_votes = {c["party"]: c["votes"] for c in entry["cands"]}
                            party_names = {c["party"]: c["name"] for c in entry["cands"]}

                            total = sum(party_votes.values())
                            # Set raw electoral votes column (column 2)
                            state_electoral_votes = sum(c.get("electoralVotes", 0) for c in entry.get("cands", []))
                            ws.cell(row=row_idx, column=2, value=state_electoral_votes)

                            sorted_parties = sorted(party_votes.items(), key=lambda x: x[1], reverse=True)
                            winner_party = sorted_parties[0][0]
                            margin = int(round(sorted_parties[0][1] - (sorted_parties[1][1] if len(sorted_parties) > 1 else 0)))
                            margin_pct = round(margin / total * 100, 2) if total else 0
                            rating = assign_rating(margin_pct, party_labels.get(winner_party, winner_party), tilt_max, lean_max, likely_max)

                            col = 3
                            for p in candidate_parties:
                                v = int(round(party_votes.get(p, 0)))
                                pct = round(v / total * 100, 2) if total else 0
                                candidate_name = party_to_candidate[p]

                                candidate_ev = next((c.get("electoralVotes", 0) for c in entry.get("cands", []) if c["name"] == candidate_name), 0)
                                display_ev = candidate_ev if p == winner_party else "-"

                                ws.cell(row=row_idx, column=col, value=f"{v:,}")                      # Raw votes
                                ws.cell(row=row_idx, column=col + 1, value=f"{pct:.2f}%")            # %
                                ws.cell(row=row_idx, column=col + 2, value=display_ev)               # Electoral votes
                                total_votes[p] += v
                                if p == winner_party:
                                    electoral_totals[p] += candidate_ev
                                col += 3

                            ws.cell(row=row_idx, column=col, value=f"{margin:,}")
                            ws.cell(row=row_idx, column=col + 1, value=f"{margin_pct:.2f}%")
                            ws.cell(row=row_idx, column=col + 2, value=f"{int(round(total)):,}")
                            ws.cell(row=row_idx, column=col + 3, value=rating)
                            row_idx += 1

                        # === Totals row ===
                        ws.cell(row=row_idx, column=1, value="TOTALS")
                        ws.cell(row=row_idx, column=2, value=f"{sum(entry.get('electoralVotes', 0) for entry in entries_to_convert)}")

                        col = 3
                        grand_total = sum(total_votes.values())
                        sorted_totals = sorted(total_votes.items(), key=lambda x: x[1], reverse=True)
                        winner_party = sorted_totals[0][0]
                        top = sorted_totals[0][1]
                        second = sorted_totals[1][1] if len(sorted_totals) > 1 else 0
                        margin = int(round(top - second))
                        margin_pct = round(margin / grand_total * 100, 2) if grand_total else 0
                        rating = assign_rating(margin_pct, party_labels.get(winner_party, winner_party), tilt_max, lean_max, likely_max)

                        for p in candidate_parties:
                            raw_total = total_votes[p]
                            ws.cell(row=row_idx, column=col, value=f"{raw_total:,}")  # Vote total
                            pct = round(raw_total / grand_total * 100, 2) if grand_total else 0
                            ws.cell(row=row_idx, column=col + 1, value=f"{pct:.2f}%")
                            ws.cell(row=row_idx, column=col + 2, value=electoral_totals[p])  # Electoral votes total
                            col += 3

                        # Calculate margins first
                        margin_total = sorted_totals[0][1] - (sorted_totals[1][1] if len(sorted_totals) >1 > 1 else 0)
                        margin_pct_total = round(margin_total / grand_total* 100, 2) if grand_total else 0

                        # Then use the calculated values
                        ws.cell(row=row_idx, column=col, value=f"{margin_total:,}")
                        ws.cell(row=row_idx, column=col + 1, value=f"{margin_pct_total:.2f}%")
                        ws.cell(row=row_idx, column=col + 2, value=f"{int(round(grand_total)):,}")
                        ws.cell(row=row_idx, column=col + 3, value=rating)

                        for c in range(1, col + 4):
                            ws.cell(row=row_idx, column=c).font = Font(bold=True)

                    # Save to buffer
                    file_stream = BytesIO()
                    wb.save(file_stream)
                    file_stream.seek(0)

                    # === DISPLAY PREVIEW ===
                    excel_rows = []
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                        excel_rows.append(list(row))

                    header_row = []
                    if len(excel_rows) >= 2:
                        row1 = excel_rows[0]
                        row2 = excel_rows[1]
                        used_names = {}

                        for col1, col2 in zip(row1, row2):
                            if col1 and col2:
                                label = f"{col1} - {col2}"
                            elif col1:
                                label = str(col1)
                            elif col2:
                                label = str(col2)
                            else:
                                label ="Unnamed"

                            if label in used_names:
                                count = used_names[label] + 1
                                used_names[label] = count
                                label = f"{label} ({count})"
                            else:
                                used_names[label] = 1

                            header_row.append(label)

                    data_rows = excel_rows[2:]
                    df_display = pd.DataFrame(data_rows, columns=header_row)
                    # Convert all numeric-like strings to numeric values
                    df_display = df_display.apply(pd.to_numeric, errors='ignore')

                    st.subheader("🧾 Presidential National View")
                    st.dataframe(df_display, use_container_width=True)

                    st.download_button(
                        label="📥 Download Presidential Spreadsheet",
                        data=file_stream,
                        file_name="Presidential_National_View.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key="president_national_view"
                    )

                    # === Presidential National View Map ===
                    pres_path = os.path.join("SVG", "presidential.svg")
                    if os.path.exists(pres_path):
                        render_svg_file(pres_path, title="🗺️ Presidential National Map", df_display=df_display, dem_colors=dem_colors, rep_colors=rep_colors, ind_colors=ind_colors)
                    else:
                        st.warning("No national map found for President.")
                # === Senate/Governor National View Spreadsheet Generator ===
                elif selected_election_type in ["Senate", "Governor"] and selected_state == "National View":
                    from collections import defaultdict
                    wb = Workbook()
                    ws = wb.active
                    ws.title = f"{selected_election_type} National View"

                    entries = election_data.get("elections", [])
                    party_labels = {"D": "Democratic", "R": "Republican", "I": "Independent"}
                    party_order = ["D", "R", "I"]
                    seats_won = {party: 0 for party in party_order}

                    # === Header Rows ===
                    ws.cell(row=2, column=1, value="State")
                    col = 2
                    for party in party_order:
                        ws.cell(row=1, column=col, value=party_labels[party])
                        ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
                        ws.cell(row=2, column=col, value="Candidate")
                        ws.cell(row=2, column=col + 1, value="#")
                        ws.cell(row=2, column=col + 2, value="%")
                        col += 3

                    ws.cell(row=1, column=col, value="Margins & Rating")
                    ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
                    ws.cell(row=2, column=col, value="Margin #")
                    ws.cell(row=2, column=col + 1, value="Margin %")
                    ws.cell(row=2, column=col + 2, value="Total Vote")
                    ws.cell(row=2, column=col + 3, value="Rating")

                    for r in range(1, 3):
                        for c in range(1, col + 4):
                            cell = ws.cell(row=r, column=c)
                            cell.font = Font(bold=True)
                            cell.alignment = Alignment(horizontal="center")

                    row_idx = 3
                    totals = {party: 0 for party in party_order}
                    grand_total = 0

                    for entry in entries:
                        state = state_code_to_name.get(entry.get("state", "???"), entry.get("state", "???"))
                        ws.cell(row=row_idx, column=1, value=state)

                        # Group candidates by party
                        party_groups = defaultdict(list)
                        for c in entry.get("cands", []):
                            party_groups[c["party"]].append(c)

                        # Determine winner
                        all_cands = sorted(entry.get("cands", []), key=lambda x: x["votes"], reverse=True)
                        winner = all_cands[0]["name"] if all_cands else None
                        winner_party = all_cands[0]["party"] if all_cands else None

                        # Prepare vote summary by party
                        party_votes = {}
                        party_names = {}
                        total_vote = sum(c["votes"] for c in entry.get("cands", []))

                        for party in party_order:
                            candidates = sorted(party_groups.get(party, []), key=lambda x: x["votes"], reverse=True)
                            if not candidates:
                                party_names[party] = ""
                                party_votes[party] = 0
                                continue

                            if len(candidates) == 1:
                                party_names[party] = candidates[0]["name"]
                                party_votes[party] = candidates[0]["votes"]
                            else:
                                if candidates[0]["name"] == winner:
                                    combined = sum(c["votes"] for c in candidates)
                                    party_names[party] = candidates[0]["name"]
                                    party_votes[party] = combined
                                else:
                                    party_names[party] = candidates[0]["name"]
                                    party_votes[party] = candidates[0]["votes"]
                                    # Move lowest-vote candidate to Independent
                                    lowest = candidates[-1]
                                    # Safely initialize Independent party
                                    if "I" not in party_votes:
                                        party_votes["I"] = 0
                                    if "I" not in party_names:
                                        party_names["I"] = ""
                                    party_names["I"] = lowest["name"]
                                    party_votes["I"] += lowest["votes"]

                        col_idx = 2
                        for party in party_order:
                            name = party_names.get(party, "")
                            votes = int(round(party_votes.get(party, 0)))
                            pct = round(votes / total_vote * 100, 2) if total_vote else 0

                            ws.cell(row=row_idx, column=col_idx, value=name)
                            ws.cell(row=row_idx, column=col_idx + 1, value=f"{votes:,}")
                            ws.cell(row=row_idx, column=col_idx + 2, value=f"{pct:.2f}%")

                            totals[party] += votes
                            col_idx += 3

                        sorted_votes = sorted(party_votes.items(), key=lambda x: x[1], reverse=True)
                        margin = int(round(sorted_votes[0][1] - (sorted_votes[1][1] if len(sorted_votes) > 1 else 0)))
                        margin_pct = round(margin / total_vote * 100, 2) if total_vote else 0
                        winner_party = sorted_votes[0][0]
                        if winner_party in seats_won:
                            seats_won[winner_party] += 1
                        rating = assign_rating(margin_pct, party_labels.get(winner_party, winner_party), tilt_max, lean_max, likely_max)

                        ws.cell(row=row_idx, column=col_idx, value=f"{margin:,}")
                        ws.cell(row=row_idx, column=col_idx + 1, value=f"{margin_pct:.2f}%")
                        ws.cell(row=row_idx, column=col_idx + 2, value=f"{int(round(total_vote)):,}")
                        ws.cell(row=row_idx, column=col_idx + 3, value=rating)

                        grand_total += total_vote
                        row_idx += 1

                    # === Totals Row ===
                    ws.cell(row=row_idx, column=1, value="TOTALS")
                    col_idx = 2
                    for party in party_order:
                        total = totals[party]
                        pct = round(total / grand_total * 100, 2) if grand_total else 0
                        ws.cell(row=row_idx, column=col_idx, value=f"{seats_won[party]} seats")
                        ws.cell(row=row_idx, column=col_idx + 1, value=f"{total:,}")
                        ws.cell(row=row_idx, column=col_idx + 2, value=f"{pct:.2f}%")
                        col_idx += 3

                    # Calculate margin for totals row
                    sorted_totals = sorted(totals.items(), key=lambda x: x[1], reverse=True)
                    margin_total = sorted_totals[0][1] - (sorted_totals[1][1] if len(sorted_totals) > 1 else 0)
                    margin_pct_total = round(margin_total / grand_total * 100, 2) if grand_total else 0
                    winner_party = sorted_totals[0][0]
                    rating = assign_rating(margin_pct_total, party_labels.get(winner_party, winner_party), tilt_max, lean_max, likely_max)

                    ws.cell(row=row_idx, column=col_idx, value=f"{margin_total:,}")
                    ws.cell(row=row_idx, column=col_idx + 1, value=f"{margin_pct_total:.2f}%")
                    ws.cell(row=row_idx, column=col_idx + 2, value=f"{int(round(grand_total)):,}")
                    ws.cell(row=row_idx, column=col_idx + 3, value=rating)

                    for c in range(1, col_idx + 4):
                        ws.cell(row=row_idx, column=c).font = Font(bold=True)

                    # === Streamlit Display ===
                    st.subheader(f"🧾 {selected_election_type} National View")
                    file_stream = BytesIO()
                    wb.save(file_stream)
                    file_stream.seek(0)

                    # === Convert Excel data to preview ===
                    excel_rows = []
                    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                        excel_rows.append(list(row))

                    header_row = []
                    if len(excel_rows) >= 2:
                        row1 = excel_rows[0]
                        row2 = excel_rows[1]
                        used_names = {}

                        for col1, col2 in zip(row1, row2):
                            if col1 and col2:
                                label = f"{col1} - {col2}"
                            elif col1:
                                label = str(col1)
                            elif col2:
                                label = str(col2)
                            else:
                                label = "Unnamed"

                            if label in used_names:
                                count = used_names[label] + 1
                                used_names[label] = count
                                label = f"{label} ({count})"
                            else:
                                used_names[label] = 1

                            header_row.append(label)

                    data_rows = excel_rows[2:]
                    df_display = pd.DataFrame(data_rows, columns=header_row)
                    # Convert all numeric-like strings to numeric values
                    df_display = df_display.apply(pd.to_numeric, errors='ignore')

                    st.dataframe(df_display, use_container_width=True)

                    st.download_button(
                        label=f"📥 Download {selected_election_type} Spreadsheet",
                        data=file_stream,
                        file_name=f"{selected_election_type.replace(' ', '_')}_National_View.xlsx",
                        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        key=f"{selected_election_type.lower().replace(' ', '_')}_national_view"
                    )



                    # === National View Maps ===
                    if selected_state == "National View":
                        if selected_election_type == "President":
                            pres_path = os.path.join("SVG", "presidential.svg")
                            if os.path.exists(pres_path):
                                render_svg_file(pres_path, title="🗺️ Presidential National Map", df_display=df_display, dem_colors=dem_colors, rep_colors=rep_colors, ind_colors=ind_colors)
                        elif selected_election_type in ["Senate", "Governor"]:
                            states_path = os.path.join("SVG", "states.svg")
                            if os.path.exists(states_path):
                                render_svg_file(states_path, title=f"🗺️ {selected_election_type} National Map", df_display=df_display, dem_colors=dem_colors, rep_colors=rep_colors, ind_colors=ind_colors)


        # === State Legislature National View Spreadsheet Generator ===
        elif selected_election_type in ["State House", "State Senate"]:

            from collections import defaultdict
            data_key = "electNightStH" if selected_election_type == "State House" else "electNightStS"
            wb = Workbook()
            ws = wb.active
            ws.title = f"{selected_election_type} National View"

            entries = sorted(election_data.get("elections", []), key=lambda x: int(x.get("district", 0)))
            party_labels = {"D": "Democratic", "R": "Republican", "I": "Independent"}
            party_order = ["D", "R", "I"]
            seats_won = {"D": 0, "R": 0, "I": 0}  # Initialize seats counter

            # === Header Rows ===
            ws.cell(row=1, column=1, value="District")
            ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
            col = 2
            for party in party_order:
                ws.cell(row=1, column=col, value=party_labels[party])
                ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 2)
                ws.cell(row=2, column=col, value="Candidate")
                ws.cell(row=2, column=col + 1, value="#")
                ws.cell(row=2, column=col + 2, value="%")
                col += 3

            ws.cell(row=1, column=col, value="Margins & Rating")
            ws.merge_cells(start_row=1, start_column=col, end_row=1, end_column=col + 3)
            ws.cell(row=2, column=col, value="Margin #")
            ws.cell(row=2, column=col + 1, value="Margin %")
            ws.cell(row=2, column=col + 2, value="Total Vote")
            ws.cell(row=2, column=col + 3, value="Rating")

            for r in range(1, 3):
                for c in range(1, col + 4):
                    cell = ws.cell(row=r, column=c)
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal="center")

            row_idx = 3
            totals= {party: 0 for party in party_order}
            grand_total = 0

            district_counter= 1
            for entry in entries:
                ws.cell(row=row_idx, column=1, value=district_counter)
                district_counter += 1

                # Group candidates by party and find winner
                party_groups = defaultdict(list)
                # Findwinner based on vote count
                candidates = sorted(entry.get("cands", []), key=lambda x: x["votes"], reverse=True)
                if candidates:
                    winner = candidates[0]["name"]
                    winner_party = candidates[0]["party"]
                    if winner_party in seats_won:
                        seats_won[winner_party] += 1

                # Group candidates by party
                for c in entry.get("cands", []):
                            party_groups[c["party"]].append(c)

                # Prepare vote summary by party
                party_votes = {}
                party_names = {}
                total_vote = sum(c["votes"] for c in entry.get("cands", []))

                for party in party_order:
                    candidates = sorted(party_groups.get(party, []), key=lambda x: x["votes"], reverse=True)
                    if not candidates:
                        party_names[party] = ""
                        party_votes[party] = 0
                        continue

                    if len(candidates) == 1:
                        party_names[party] = candidates[0]["name"]
                        party_votes[party] = candidates[0]["votes"]
                    else:
                        winner_in_group = next((c for c in candidates if c["name"] == winner), None)
                        if winner_in_group:
                            combined = sum(c["votes"] for c in candidates)
                            party_names[party]= candidates[0]["name"]
                            party_votes[party] = combined
                        else:
                            party_names[party] = candidates[0]["name"]
                            party_votes[party] = candidates[0]["votes"]
                            # Move lowest-vote candidate to Independent
                            lowest = candidates[-1]
                            # Safely initialize Independent party
                            if "I" not in party_votes:
                                party_votes["I"] = 0
                            if "I" not in party_names:
                                party_names["I"] = ""
                            party_names["I"] = lowest["name"]
                            party_votes["I"] += lowest["votes"]

                col_idx = 2
                for party in party_order:
                    name = party_names.get(party, "")
                    votes = int(round(party_votes.get(party, 0)))
                    pct = round(votes / total_vote * 100, 2) if total_vote else 0

                    ws.cell(row=row_idx, column=col_idx, value=name)
                    ws.cell(row=row_idx, column=col_idx + 1, value=f"{votes:,}")
                    ws.cell(row=row_idx, column=col_idx + 2, value=f"{pct:.2f}%")

                    totals[party] += votes
                    col_idx += 3

                sorted_votes = sorted(party_votes.items(), key=lambda x: x[1], reverse=True)
                margin = int(round(sorted_votes[0][1] - (sorted_votes[1][1] if len(sorted_votes) > 1 else 0)))
                margin_pct = round(margin / total_vote * 100, 2) if total_vote else 0
                rating = assign_rating(margin_pct, party_labels.get(sorted_votes[0][0], sorted_votes[0][0]), st.session_state["tilt_max"], st.session_state["lean_max"], st.session_state["likely_max"])

                ws.cell(row=row_idx, column=col_idx, value=f"{margin:,}")
                ws.cell(row=row_idx, column=col_idx + 1, value=f"{margin_pct:.2f}%")
                ws.cell(row=row_idx, column=col_idx + 2, value=f"{int(round(total_vote)):,}")
                ws.cell(row=row_idx, column=col_idx + 3, value=rating)

                grand_total += total_vote
                row_idx += 1

            # === Totals Row ===
            ws.cell(row=row_idx, column=1, value="TOTALS")
            col_idx = 2
            for party in party_order:
                total = totals[party]
                pct = round(total / grand_total * 100, 2) if grand_total else 0
                seats = seats_won[party]
                ws.cell(row=row_idx, column=col_idx, value=f"{seats} seats")
                ws.cell(row=row_idx, column=col_idx + 1, value=f"{total:,}")
                ws.cell(row=row_idx, column=col_idx + 2, value=f"{pct:.2f}%")
                col_idx += 3

            sorted_totals = sorted(totals.items(), key=lambda x: x[1], reverse=True)
            margin_total = sorted_totals[0][1] - (sorted_totals[1][1] if len(sorted_totals) > 1 else 0)
            margin_pct_total = round(margin_total / grand_total * 100, 2) if grand_total else 0
            winner_party = sorted_totals[0][0]
            rating = assign_rating(margin_pct_total, party_labels.get(winner_party, winner_party), st.session_state["tilt_max"], st.session_state["lean_max"], st.session_state["likely_max"])

            ws.cell(row=row_idx, column=col_idx, value=f"{margin_total:,}")
            ws.cell(row=row_idx, column=col_idx + 1, value=f"{margin_pct_total:.2f}%")
            ws.cell(row=row_idx, column=col_idx + 2, value=f"{int(round(grand_total)):,}")
            ws.cell(row=row_idx, column=col_idx + 3, value=rating)

            for c in range(1, col_idx + 4):
                ws.cell(row=row_idx, column=c).font = Font(bold=True)

            # === Streamlit Display ===
            st.subheader(f"🧾 {selected_election_type} National View")
            file_stream = BytesIO()
            wb.save(file_stream)
            file_stream.seek(0)

            # Convert worksheet to displayable rows
            excel_rows = []
            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, values_only=True):
                excel_rows.append(list(row))

            header_row = []
            if len(excel_rows) >= 2:
                row1 = excel_rows[0]
                row2 = excel_rows[1]
                used_names = {}

                for col1, col2 in zip(row1, row2):
                    if col1 and col2:
                        label = f"{col1} - {col2}"
                    elif col1:
                        label = str(col1)
                    elif col2:
                        label = str(col2)
                    else:
                        label = "Unnamed"

                    if label in used_names:
                        count = used_names[label] + 1
                        used_names[label] = count
                        label = f"{label} ({count})"
                    else:
                        used_names[label] = 1

                    header_row.append(label)

            data_rows = excel_rows[2:]

            if header_row and data_rows:
                df_display = pd.DataFrame(data_rows, columns=header_row)
                # Convert all numeric-like strings to numeric values
                df_display = df_display.apply(pd.to_numeric, errors='ignore')
                # Apply custom ratings based on session state thresholds
                df_display = update_df_with_custom_ratings(
                    df_display
                )
                st.dataframe(df_display, use_container_width=True)

            # Download button
            st.download_button(
                label=f"📥 Download {selected_election_type} Spreadsheet",
                data=file_stream,
                file_name=f"{selected_election_type.replace(' ', '_')}_National_View.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                key=f"{selected_election_type.lower().replace(' ', '_')}_national_view"
            )
        else:
            st.warning("This election type is not yet supported.")
    else:
        st.warning("No recognized election data found in this file.")